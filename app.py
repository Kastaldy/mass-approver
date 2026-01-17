from flask import Flask, render_template, request, send_file, jsonify, send_from_directory
import pandas as pd
import numpy as np
from io import BytesIO
import os
import tempfile
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

# === APENAS UMA INICIALIZAÇÃO DO FLASK ===
app = Flask(__name__,
            static_folder='static',
            template_folder='templates')

app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size
app.config['SECRET_KEY'] = os.urandom(24)

# Debug - mostra onde está procurando
print("=" * 60)
print(" INICIANDO ANALISADOR DE INDICADORES")
print("=" * 60)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, 'static')
TEMPLATE_DIR = os.path.join(BASE_DIR, 'templates')

print(f" Diretório base: {BASE_DIR}")
print(f" Pasta static: {STATIC_DIR}")
print(f" Pasta templates: {TEMPLATE_DIR}")
print("-" * 60)

# Verifica se os arquivos existem
arquivos_necessarios = {
    'index.html': os.path.join(TEMPLATE_DIR, 'index.html'),
    'style.css': os.path.join(STATIC_DIR, 'css', 'style.css'),
    'script.js': os.path.join(STATIC_DIR, 'js', 'script.js')
}

for nome, caminho in arquivos_necessarios.items():
    existe = os.path.exists(caminho)
    status = " ENCONTRADO" if existe else " NÃO ENCONTRADO"
    print(f"{status}: {nome}")
    if not existe and 'html' not in nome:
        print(f"   → Caminho: {caminho}")

print("=" * 60)

def carregar_e_transpor_dados(file_path):
    """Carrega e transpõe os dados automaticamente."""
    try:
        if file_path.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file_path, header=None)
        elif file_path.endswith('.csv'):
            df = pd.read_csv(file_path, header=None, encoding='utf-8')
        else:
            return None
        
        if df.shape[0] < 2 or df.shape[1] < 2:
            return None
        
        enderecos = df.iloc[0, 1:].tolist()
        indicadores = df.iloc[1:, 0].tolist()
        dados = df.iloc[1:, 1:].values
        
        df_final = pd.DataFrame(dados, index=indicadores, columns=enderecos)
        return df_final
        
    except Exception as e:
        print(f"Erro ao carregar dados: {e}")
        return None

def extrair_indicadores_principais(df):
    """Extrai os 3 indicadores principais."""
    if df is None or df.empty:
        return None
    
    resultados = {}
    
    for endereco in df.columns:
        try:
            renda_valor = df.loc['Renda média domiciliar', endereco]
            pea_valor = df.loc['PEA Dia', endereco]
            densidade_valor = df.loc['Densidade demográfica', endereco]
            
            renda = float(renda_valor)
            pea = float(pea_valor)
            densidade = float(densidade_valor)
            
            resultados[endereco] = {
                'renda_media': renda,
                'pea_dia': pea,
                'densidade': densidade
            }
        except Exception as e:
            print(f"    Erro no endereço {endereco[:30]}...: {str(e)[:50]}")
            continue
    
    if resultados:
        df_resultados = pd.DataFrame.from_dict(resultados, orient='index')
        df_resultados.index.name = 'Endereço'
        print(f" Extraídos {len(resultados)} conjuntos de indicadores válidos")
        return df_resultados
    
    return None

def analisar_indicadores(df_indicadores, regua_renda, regua_pea, regua_densidade):
    """Analisa os indicadores e verifica se estão acima das réguas."""
    if df_indicadores is None or df_indicadores.empty:
        return None
    
    resultados = []
    
    for endereco, dados in df_indicadores.iterrows():
        # Converter valores numpy para tipos Python padrão
        renda = float(dados['renda_media'])  # Garante float do Python
        pea = float(dados['pea_dia'])        # Garante float do Python
        densidade = float(dados['densidade']) # Garante float do Python
        
        acima_renda = renda >= regua_renda
        acima_pea = pea >= regua_pea
        acima_densidade = densidade >= regua_densidade
        
        contador = sum([acima_renda, acima_pea, acima_densidade])
        
        if contador == 3:
            status = "APROVADO"
            status_class = "aprovado"
        elif contador == 2:
            status = "PARCIAL (2/3)"
            status_class = "parcial"
        else:
            status = "REPROVADO"
            status_class = "reprovado"
        
        resultados.append({
            'endereco': str(endereco),  # Garante string
            'renda_media': float(renda),  # Garante float Python
            'pea_dia': float(pea),        # Garante float Python
            'densidade': float(densidade), # Garante float Python
            'acima_renda': '✓' if acima_renda else '✗',
            'acima_pea': '✓' if acima_pea else '✗',
            'acima_densidade': '✓' if acima_densidade else '✗',
            'pontos': int(contador),      # Garante int Python
            'status': status,
            'status_class': status_class
        })
    
    return resultados

def to_excel(resultados):
    """Converte resultados para Excel."""
    if not resultados:
        return None
    
    # Criar DataFrame
    df = pd.DataFrame(resultados)
    
    # Renomear colunas para melhor visualização
    df = df.rename(columns={
        'endereco': 'Endereço',
        'renda_media': 'Renda Média (R$)',
        'pea_dia': 'PEA Dia',
        'densidade': 'Densidade Demográfica',
        'acima_renda': 'Renda OK',
        'acima_pea': 'PEA OK',
        'acima_densidade': 'Densidade OK',
        'pontos': 'Pontos Acima',
        'status': 'Status',
        'status_class': 'Classe Status'
    })
    
    output = BytesIO()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    
    # Escrever cabeçalho
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Escrever dados
    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Aplicar formatação condicional
    for row in range(2, len(df) + 2):
        status_cell = ws.cell(row=row, column=headers.index('Status') + 1)
        status_value = status_cell.value
        
        if status_value == 'APROVADO':
            fill_color = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif 'PARCIAL' in str(status_value):
            fill_color = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif status_value == 'REPROVADO':
            fill_color = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            fill_color = None
        
        if fill_color:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill_color
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output

# ========== ROTAS ==========
@app.route('/')
def index():
    """Página principal."""
    return render_template('index.html')

@app.route('/analisar', methods=['POST'])
def analisar():
    """Endpoint para análise dos dados."""
    try:
        if 'file' not in request.files:
            return jsonify({'erro': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'erro': 'Nenhum arquivo selecionado'}), 400
        
        # Obter parâmetros
        regua_renda = float(request.form.get('regua_renda', 4600))
        regua_pea = float(request.form.get('regua_pea', 5000))
        regua_densidade = float(request.form.get('regua_densidade', 5000))
        
        print(f" Análise iniciada - Réguas: R${regua_renda}, PEA:{regua_pea}, Dens:{regua_densidade}")
        print(f" Arquivo: {file.filename}")
        
        # Salvar arquivo temporariamente
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as tmp:
            file.save(tmp.name)
            file_path = tmp.name
        
        try:
            # Processar arquivo
            df_original = carregar_e_transpor_dados(file_path)
            
            if df_original is None:
                return jsonify({'erro': 'Erro ao processar arquivo. Verifique o formato.'}), 400
            
            # Extrair indicadores
            df_indicadores = extrair_indicadores_principais(df_original)
            
            if df_indicadores is None:
                return jsonify({'erro': 'Não foi possível extrair os indicadores do arquivo.'}), 400
            
            # Analisar
            resultados = analisar_indicadores(df_indicadores, regua_renda, regua_pea, regua_densidade)
            
            if resultados:
                # Calcular estatísticas
                total = len(resultados)
                aprovados = len([r for r in resultados if r['status'] == 'APROVADO'])
                parciais = len([r for r in resultados if 'PARCIAL' in r['status']])
                reprovados = len([r for r in resultados if r['status'] == 'REPROVADO'])
                
                print(f" Análise concluída: {total} endereços processados")
                print(f"   Aprovados: {aprovados}, Parciais: {parciais}, Reprovados: {reprovados}")
                
                return jsonify({
                    'sucesso': True,
                    'resultados': resultados,
                    'estatisticas': {
                        'total': total,
                        'aprovados': aprovados,
                        'parciais': parciais,
                        'reprovados': reprovados,
                        'taxa_aprovacao': f"{(aprovados/total*100):.1f}%" if total > 0 else "0%",
                        'taxa_parciais': f"{(parciais/total*100):.1f}%" if total > 0 else "0%",
                        'taxa_reprovacao': f"{(reprovados/total*100):.1f}%" if total > 0 else "0%"
                    }
                })
            else:
                return jsonify({'erro': 'Nenhum resultado encontrado.'}), 400
                
        finally:
            # Limpar arquivo temporário
            if os.path.exists(file_path):
                os.unlink(file_path)
            
    except Exception as e:
        print(f" Erro no processamento: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

@app.route('/download', methods=['POST'])
def download():
    """Endpoint para download do Excel."""
    try:
        data = request.get_json()
        resultados = data.get('resultados', [])
        
        if not resultados:
            return jsonify({'erro': 'Nenhum dado para exportar'}), 400
        
        print(f" Gerando Excel com {len(resultados)} resultados...")
        
        # Gerar Excel
        excel_file = to_excel(resultados)
        
        if excel_file is None:
            return jsonify({'erro': 'Erro ao gerar arquivo Excel'}), 500
        
        # Nome do arquivo com data
        data_hora = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'resultados_analise_{data_hora}.xlsx'
        
        print(f" Excel gerado: {filename}")
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f" Erro no download: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'erro': f'Erro ao gerar arquivo: {str(e)}'}), 500

# Rota para arquivos estáticos (IMPORTANTE!)
@app.route('/static/<path:filename>')
def serve_static(filename):
    """Serve arquivos estáticos."""
    return send_from_directory('static', filename)

# Rota de teste
@app.route('/teste')
def teste():
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Teste</title>
        <style>
            body { font-family: Arial; padding: 20px; }
            .teste { background: #f0f0f0; padding: 20px; margin: 10px; }
        </style>
    </head>
    <body>
        <h1>✅ Flask está funcionando!</h1>
        <div class="teste">
            <h3>Teste de funcionalidades:</h3>
            <p><a href="/">Voltar para página principal</a></p>
            <p><a href="/static/css/style.css">Testar CSS</a></p>
            <p><a href="/static/js/script.js">Testar JavaScript</a></p>
        </div>
    </body>
    </html>
    '''

# Rota para debug
@app.route('/debug')
def debug():
    """Página de debug para verificar estrutura."""
    info = {
        'diretorio_atual': os.getcwd(),
        'app_dir': BASE_DIR,
        'static_dir': STATIC_DIR,
        'template_dir': TEMPLATE_DIR,
        'arquivos': {}
    }
    
    for nome, caminho in arquivos_necessarios.items():
        info['arquivos'][nome] = {
            'caminho': caminho,
            'existe': os.path.exists(caminho),
            'tamanho': os.path.getsize(caminho) if os.path.exists(caminho) else 0
        }
    
    html = f'''
    <!DOCTYPE html>
    <html>
    <head><title>Debug</title></head>
    <body>
        <h1> Informações de Debug</h1>
        <h2>Diretórios:</h2>
        <pre>{info}</pre>
        <h2>Arquivos necessários:</h2>
        <ul>
    '''
    
    for nome, dados in info['arquivos'].items():
        status = "Positivo - " if dados['existe'] else "Negativo - "
        html += f'<li>{status} {nome}: {dados["caminho"]}</li>'
    
    html += '''
        </ul>
        <p><a href="/">← Voltar</a></p>
    </body>
    </html>
    '''
    
    return html


if __name__ == '__main__':
    import sys
    print(f"\n✅ Python {sys.version}")
    
    port = int(os.environ.get("PORT", 5000))
    
    print("\n" + "=" * 60)
    print("🚀 Mass Approver iniciando...")
    print(f"🔧 Porta: {port}")
    print("=" * 60 + "\n")
    
    app.run(debug=False, host='0.0.0.0', port=port)
